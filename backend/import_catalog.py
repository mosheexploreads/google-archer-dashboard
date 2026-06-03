"""
Import Archer catalog from Excel into product_catalog table.
Run from backend/ directory: python import_catalog.py
"""
import openpyxl
from datetime import datetime
from app.database import SessionLocal, engine, Base
from app.models import ProductCatalog

def import_catalog():
    """Read Excel and insert into product_catalog."""
    excel_path = r'C:\Users\moshe\Downloads\Archer\Archer Catalog 2026-06-03.xlsx'

    # Create tables if they don't exist
    Base.metadata.create_all(bind=engine)
    print("Database tables created/verified.")

    # Load workbook
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Build header map
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers, 1)}  # 1-indexed

    print(f"Found {len(headers)} columns")
    print(f"Total rows to process: {ws.max_row - 1}")

    db = SessionLocal()

    try:
        # Clear existing data
        db.query(ProductCatalog).delete()
        db.commit()
        print("Cleared existing product_catalog.\n")

        inserted = 0
        skipped_dups = 0
        errors = 0

        # Process each row with proper error handling
        for row_idx in range(2, ws.max_row + 1):
            try:
                asin = ws.cell(row_idx, col_map['ASIN']).value
                if not asin:
                    continue

                asin_upper = str(asin).upper()

                # Check for duplicate before querying
                existing = db.query(ProductCatalog).filter(
                    ProductCatalog.asin == asin_upper,
                    ProductCatalog.country_code == 'US'
                ).first()

                if existing:
                    skipped_dups += 1
                    continue

                # Extract fields
                product_name = ws.cell(row_idx, col_map['Product Name']).value
                rating_val = ws.cell(row_idx, col_map['Avg Rating']).value
                review_count_val = ws.cell(row_idx, col_map['Total Reviews']).value
                price_val = ws.cell(row_idx, col_map['Final Price']).value
                image_url = ws.cell(row_idx, col_map['Image URL']).value
                availability = ws.cell(row_idx, col_map['Product Status']).value

                # Convert types safely
                try:
                    rating = float(rating_val) if rating_val else None
                except:
                    rating = None

                try:
                    review_count = int(review_count_val) if review_count_val else 0
                except:
                    review_count = 0

                try:
                    price = float(price_val) if price_val else None
                except:
                    price = None

                # Create and add row
                catalog_row = ProductCatalog(
                    asin=asin_upper,
                    country_code='US',
                    product_name=product_name,
                    price=price,
                    rating=rating,
                    review_count=review_count,
                    image_url=image_url,
                    availability=availability,
                    last_synced_at=datetime.utcnow(),
                )
                db.add(catalog_row)
                inserted += 1

                # Commit periodically
                if inserted % 10000 == 0:
                    db.commit()
                    print(f"  Inserted {inserted} rows, skipped {skipped_dups} duplicates...")

            except Exception as e:
                # Roll back this transaction and continue
                db.rollback()
                errors += 1
                if errors <= 5:  # Only print first 5 errors
                    print(f"  Error on row {row_idx}: {e}")
                continue

        # Final commit
        db.commit()
        print(f"\nImport complete!")
        print(f"  Inserted: {inserted}")
        print(f"  Skipped (duplicates): {skipped_dups}")
        print(f"  Errors: {errors}")

        # Show stats
        total = db.query(ProductCatalog).count()
        with_rating = db.query(ProductCatalog).filter(ProductCatalog.rating.isnot(None)).count()
        with_reviews = db.query(ProductCatalog).filter(ProductCatalog.review_count > 0).count()
        matching = db.query(ProductCatalog).filter(
            ProductCatalog.rating >= 4.2,
            ProductCatalog.review_count >= 100
        ).count()

        print(f"\nDatabase stats:")
        print(f"  Total products: {total}")
        print(f"  With rating: {with_rating}")
        print(f"  With reviews: {with_reviews}")
        print(f"  Matching (4.2+ rating AND 100+ reviews): {matching}")

    finally:
        db.close()

if __name__ == '__main__':
    import_catalog()
